from Utils.tools import Tools, CustomException
from Utils.querys import Querys
import pandas as pd
import base64
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class Contador:
    """
    Clase para gestionar la lógica de negocio de años y objetivos del plan de ventas.
    """

    def __init__(self, db):
        self.db = db
        self.tools = Tools()
        self.querys = Querys(self.db)

    # Función para procesar archivo Excel
    def procesar_archivo_excel(self, data: dict):
        """
        Procesa un archivo Excel de la DIAN aplicando filtros y cálculos.
        
        Args:
            data (dict): {
                "archivo": "base64_string",
                "nombre_archivo": "archivo.xlsx",
                "tipo_archivo": "ventas"
            }
        """
        try:
            archivo_base64 = data.get("archivo")
            nombre_archivo = data.get("nombre_archivo")
            tipo_archivo = data.get("tipo_archivo")

            if not archivo_base64 or not nombre_archivo or not tipo_archivo:
                raise CustomException("Faltan campos requeridos: archivo, nombre_archivo, tipo_archivo")

            # Decodificar el archivo Base64
            archivo_bytes = base64.b64decode(archivo_base64)
            
            # Leer el archivo Excel
            df = pd.read_excel(BytesIO(archivo_bytes))

            # Validar que las columnas necesarias existan
            columnas_requeridas = [
                'Tipo de documento', 'CUFE/CUDE', 'Folio', 'Prefijo', 'Divisa',
                'Forma de Pago', 'Medio de Pago', 'Fecha Emisión', 'Fecha Recepción',
                'NIT Emisor', 'Nombre Emisor', 'NIT Receptor', 'Nombre Receptor',
                'IVA', 'ICA', 'IC', 'INC', 'Timbre', 'INC Bolsas', 'IN Carbono',
                'IN Combustibles', 'IC Datos', 'ICL', 'INPP', 'IBUA', 'ICUI',
                'Rete IVA', 'Rete Renta', 'Rete ICA', 'Total', 'Estado', 'Grupo'
            ]

            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            if columnas_faltantes:
                raise CustomException(f"El archivo no contiene las columnas requeridas: {', '.join(columnas_faltantes)}")

            # PASO 1: Filtrar por tipo de documento
            tipos_documentos_validos = [
                "Factura electrónica",
                "Factura electrónica de contingencia",
                "Nota de crédito electrónica"
            ]
            df_filtrado = df[df['Tipo de documento'].isin(tipos_documentos_validos)]

            if df_filtrado.empty:
                raise CustomException("No se encontraron registros con los tipos de documento válidos")

            # PASO 2: Filtrar por NIT Emisor = 890101977
            df_filtrado = df_filtrado[df_filtrado['NIT Emisor'] == 890101977]

            if df_filtrado.empty:
                raise CustomException("No se encontraron registros con el NIT Emisor 890101977")

            # PASO 3: Agregar columna con fórmula: Prefijo + Folio
            # Lógica: SI(Prefijo="CRD","FC",SI(Prefijo="DV","DV",0)) & " " & Folio
            def calcular_tipo_folio(row):
                prefijo = str(row['Prefijo']).strip()
                folio = str(row['Folio']).strip()
                
                if prefijo == "CRD":
                    tipo = "FC"
                elif prefijo == "DV":
                    tipo = "DV"
                else:
                    tipo = "0"
                
                return f"{tipo} {folio}"

            df_filtrado['Tipo-Folio'] = df_filtrado.apply(calcular_tipo_folio, axis=1)

            # PASO 4: Agregar columna Subtotal = Total - IVA
            df_filtrado['Subtotal'] = df_filtrado['Total'] - df_filtrado['IVA']

            # PASO 5: Agregar columna Naturaleza de la operación
            # Lógica: SI(Prefijo="CRD",Subtotal*1,SI(Prefijo="DV",Subtotal*-1,0))
            def calcular_naturaleza_operacion(row):
                prefijo = str(row['Prefijo']).strip()
                subtotal = row['Subtotal']
                
                if prefijo == "CRD":
                    return subtotal * 1
                elif prefijo == "DV":
                    return subtotal * -1
                else:
                    return 0

            df_filtrado['Saldo2'] = df_filtrado.apply(calcular_naturaleza_operacion, axis=1)

            # Generar archivo Excel procesado con formato
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_filtrado.to_excel(writer, index=False, sheet_name='Datos Procesados')
                
                # Obtener el workbook y la hoja activa
                workbook = writer.book
                worksheet = writer.sheets['Datos Procesados']
                
                # Definir estilos para el encabezado
                header_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Verde
                header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # Blanco y negrita
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                border_side = Side(style='thin', color='000000')
                header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
                
                # Aplicar estilos a la fila de encabezado
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = header_border
                
                # Ajustar el ancho de las columnas automáticamente
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            archivo_procesado_base64 = base64.b64encode(output.read()).decode('utf-8')

            # Generar nombre del archivo de salida
            nombre_base = nombre_archivo.rsplit('.', 1)[0]
            nombre_archivo_procesado = f"{nombre_base}_procesado.xlsx"

            # Preparar datos para guardar en BD (convertir DataFrame a lista de diccionarios)
            registros_para_bd = df_filtrado.to_dict('records')
            
            # Convertir valores NaN a None para que sean serializables a JSON
            for registro in registros_para_bd:
                for key, value in registro.items():
                    if pd.isna(value):
                        registro[key] = None

            datos_para_bd = {
                "total_registros_originales": len(df),
                "total_registros_filtrados": len(df_filtrado),
                "nombre_archivo_original": nombre_archivo,
                "fecha_procesamiento": pd.Timestamp.now().isoformat(),
                "registros": registros_para_bd
            }

            # Desactivar registros anteriores del mismo tipo (tipo=1 para DIAN)
            self.querys.desactivar_registros_anteriores(tipo=1)
            
            # Guardar nuevos datos procesados
            self.querys.guardar_datos_procesados(tipo=1, datos=datos_para_bd)

            resultado = {
                "total_registros_originales": len(df),
                "total_registros_filtrados": len(df_filtrado),
                "tipo_archivo": tipo_archivo,
                "nombre_archivo_original": nombre_archivo,
                "nombre_archivo_procesado": nombre_archivo_procesado,
                "archivo_procesado": archivo_procesado_base64
            }

            return self.tools.output(200, f"Archivo procesado exitosamente. {len(df_filtrado)} registros procesados de {len(df)} originales.", resultado)
            
        except CustomException as e:
            raise e
        except Exception as e:
            print(f"Error al procesar archivo Excel: {e}")
            raise CustomException(f"Error al procesar archivo Excel: {str(e)}")

    # Función para procesar archivo DMS
    def procesar_archivo_dms(self, data: dict):
        """
        Procesa un archivo DMS aplicando cálculos.
        
        Args:
            data (dict): {
                "archivo": "base64_string",
                "nombre_archivo": "archivo.xlsx",
                "tipo_archivo": "dms"
            }
        """
        try:
            archivo_base64 = data.get("archivo")
            nombre_archivo = data.get("nombre_archivo")
            tipo_archivo = data.get("tipo_archivo")

            if not archivo_base64 or not nombre_archivo or not tipo_archivo:
                raise CustomException("Faltan campos requeridos: archivo, nombre_archivo, tipo_archivo")

            # Decodificar el archivo Base64
            archivo_bytes = base64.b64decode(archivo_base64)
            
            # Leer el archivo Excel
            df = pd.read_excel(BytesIO(archivo_bytes))

            # Validar que las columnas necesarias existan
            columnas_requeridas = [
                'Cuenta Nivel 10', 'Descripción Cuenta', 'Tipo Docto.', 'Descripción Tipo',
                'Número Docto.', 'Mes Docto.', 'Fecha Docto.', 'Tercero', 'Nombre Tercero',
                'Centro de Costo', 'Descripción Centro', 'Débito', 'Crédito', 'Saldo Periodo',
                'Base', 'Débito Niif', 'Crédito Niif', 'Saldo Periodo Niif', 'Explicación'
            ]

            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            if columnas_faltantes:
                raise CustomException(f"El archivo no contiene las columnas requeridas: {', '.join(columnas_faltantes)}")

            # PASO 1: Agregar columna tipo_doc_desc_tipo = Tipo Docto. + " " + Número Docto.
            df['tipo_doc_desc_tipo'] = df['Tipo Docto.'].astype(str) + " " + df['Número Docto.'].astype(str)

            # PASO 2: Agregar columna Saldo2
            # Lógica: SI(Tipo Docto.="FC", Saldo Periodo*-1, SI(Tipo Docto.="DV", Saldo Periodo*-1, 0))
            def calcular_saldo2(row):
                tipo_docto = str(row['Tipo Docto.']).strip()
                saldo_periodo = row['Saldo Periodo']
                
                if tipo_docto == "FC":
                    return saldo_periodo * -1
                elif tipo_docto == "DV":
                    return saldo_periodo * -1
                else:
                    return 0

            df['Saldo2'] = df.apply(calcular_saldo2, axis=1)

            # Generar archivo Excel procesado con formato
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Datos Procesados')
                
                # Obtener el workbook y la hoja activa
                workbook = writer.book
                worksheet = writer.sheets['Datos Procesados']
                
                # Definir estilos para el encabezado
                header_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Verde
                header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # Blanco y negrita
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                border_side = Side(style='thin', color='000000')
                header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
                
                # Aplicar estilos a la fila de encabezado
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = header_border
                
                # Ajustar el ancho de las columnas automáticamente
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            archivo_procesado_base64 = base64.b64encode(output.read()).decode('utf-8')

            # Generar nombre del archivo de salida
            nombre_base = nombre_archivo.rsplit('.', 1)[0]
            nombre_archivo_procesado = f"{nombre_base}_procesado.xlsx"

            # Preparar datos para guardar en BD (convertir DataFrame a lista de diccionarios)
            registros_para_bd = df.to_dict('records')
            
            # Convertir valores NaN a None para que sean serializables a JSON
            for registro in registros_para_bd:
                for key, value in registro.items():
                    if pd.isna(value):
                        registro[key] = None

            datos_para_bd = {
                "total_registros": len(df),
                "nombre_archivo_original": nombre_archivo,
                "fecha_procesamiento": pd.Timestamp.now().isoformat(),
                "registros": registros_para_bd
            }

            # Desactivar registros anteriores del mismo tipo (tipo=2 para DMS)
            self.querys.desactivar_registros_anteriores(tipo=2)
            
            # Guardar nuevos datos procesados
            self.querys.guardar_datos_procesados(tipo=2, datos=datos_para_bd)

            resultado = {
                "total_registros": len(df),
                "tipo_archivo": tipo_archivo,
                "nombre_archivo_original": nombre_archivo,
                "nombre_archivo_procesado": nombre_archivo_procesado,
                "archivo_procesado": archivo_procesado_base64
            }

            return self.tools.output(200, f"Archivo DMS procesado exitosamente. {len(df)} registros procesados.", resultado)
            
        except CustomException as e:
            raise e
        except Exception as e:
            print(f"Error al procesar archivo DMS: {e}")
            raise CustomException(f"Error al procesar archivo DMS: {str(e)}")
